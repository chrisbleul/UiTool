import datetime
import json

import openpyxl
import pytest

from uiflow.excel import read_excel_rows, write_excel_rows


@pytest.fixture
def workbook_path(tmp_path):
    path = tmp_path / "demo.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "betrag"])
    ws.append(["Anna", 10])
    ws.append(["Bert", 25.5])
    wb.save(path)
    return path


def test_read_excel_rows_uses_first_row_as_headers(workbook_path):
    rows = read_excel_rows(str(workbook_path))
    assert rows == [
        {"name": "Anna", "betrag": 10},
        {"name": "Bert", "betrag": 25.5},
    ]


def test_read_excel_rows_skips_fully_empty_rows(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active
    ws.append([None, None])
    ws.append(["Chris", 5])
    wb.save(workbook_path)

    rows = read_excel_rows(str(workbook_path))

    assert [r["name"] for r in rows] == ["Anna", "Bert", "Chris"]


def test_read_excel_rows_missing_file_raises():
    with pytest.raises(FileNotFoundError):
        read_excel_rows("does_not_exist.xlsx")


def test_read_excel_rows_dates_are_json_serializable(tmp_path):
    path = tmp_path / "dates.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["name", "due"])
    ws.append(["Anna", datetime.date(2026, 1, 15)])
    wb.save(path)

    rows = read_excel_rows(str(path))

    json.dumps(rows)  # must not raise
    assert isinstance(rows[0]["due"], str)
    assert rows[0]["due"].startswith("2026-01-15")


def test_write_excel_rows_writes_dict_rows_with_union_of_keys(tmp_path):
    path = tmp_path / "out.xlsx"
    rows = [{"name": "Anna", "betrag": 10}, {"name": "Bert", "betrag": 25.5, "extra": "x"}]

    count = write_excel_rows(str(path), rows)

    assert count == 2
    read_back = read_excel_rows(str(path))
    assert read_back == [
        {"name": "Anna", "betrag": 10, "extra": None},
        {"name": "Bert", "betrag": 25.5, "extra": "x"},
    ]


def test_write_excel_rows_writes_list_rows_without_header(tmp_path):
    path = tmp_path / "out.xlsx"

    write_excel_rows(str(path), [["a", 1], ["b", 2]])

    wb = openpyxl.load_workbook(path)
    ws = wb.active
    assert [list(row) for row in ws.iter_rows(values_only=True)] == [["a", 1], ["b", 2]]


def test_write_excel_rows_empty_list_creates_empty_workbook(tmp_path):
    path = tmp_path / "out.xlsx"

    count = write_excel_rows(str(path), [])

    assert count == 0
    assert path.exists()


def test_write_excel_rows_respects_sheet_name(tmp_path):
    path = tmp_path / "out.xlsx"

    write_excel_rows(str(path), [{"x": 1}], sheet="MySheet")

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["MySheet"]
