"""Reads .xlsx files into plain list-of-dict rows (a lightweight "DataTable"):
first row = column headers, every following row = one dict keyed by them.
Used both by the `read_excel` engine action and the Studio's queue-import
endpoint (each row becomes one queue item's payload)."""

from __future__ import annotations

import datetime
from pathlib import Path
from typing import Any


def _json_safe(value: Any) -> Any:
    """openpyxl returns native datetime/date/time objects for date-formatted
    cells, which json.dumps() (used when a row becomes a queue item's payload,
    or a workflow variable gets logged) can't serialize. Normalize to ISO
    strings; everything else (str/int/float/bool/None) already round-trips."""
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return value


def read_excel_rows(path: str, sheet: str | None = None) -> list[dict[str, Any]]:
    import openpyxl

    file_path = Path(path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    # read_only=True keeps a lazy handle open on the underlying zip file for
    # streaming row-by-row access; without an explicit close() that handle
    # outlives this function (until GC gets to it), which on Windows blocks a
    # caller from deleting/replacing the file right after (e.g. the Studio's
    # Excel-import endpoint, which reads into a temp file then unlinks it).
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.active

        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(next(rows_iter))]
        except StopIteration:
            return []

        rows: list[dict[str, Any]] = []
        for raw_row in rows_iter:
            if raw_row is None or all(cell is None for cell in raw_row):
                continue
            row = {headers[i]: _json_safe(raw_row[i]) for i in range(min(len(headers), len(raw_row)))}
            rows.append(row)
        return rows
    finally:
        workbook.close()


def write_excel_rows(path: str, rows: list[dict[str, Any]] | list[list[Any]], sheet: str | None = None) -> int:
    """Writes rows to a new .xlsx file at `path` (always creates/overwrites -
    there's no append-to-existing-workbook mode in this MVP). Dict rows use the
    union of keys (in first-seen order) as the header row; list rows are
    written as-is with no header."""
    import openpyxl

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    if sheet:
        worksheet.title = sheet

    if not rows:
        workbook.save(path)
        return 0

    if isinstance(rows[0], dict):
        headers: list[str] = []
        for row in rows:
            for key in row:
                if key not in headers:
                    headers.append(key)
        worksheet.append(headers)
        for row in rows:
            worksheet.append([row.get(h) for h in headers])
    else:
        for row in rows:
            worksheet.append(list(row))

    workbook.save(path)
    return len(rows)
